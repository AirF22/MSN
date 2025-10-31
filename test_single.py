import os,yaml
import cv2
import pandas as pd
import numpy as np
import torch
import time
from PIL import Image
from tqdm import tqdm
from tools import getMAE_numpy_2D
import model
from openpyxl import load_workbook, Workbook


AUTO_CALL = True  # 函数是否被自动调用
eps = 1e-10
class Tester:
    def __init__(self, log_path, config_file_name):
        with open(os.path.join(log_path, config_file_name), 'r') as f:
            self.hyper = yaml.safe_load(f)
        self.log_path = log_path  # 日志目录
        self.device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu') # 创建运算设备
        if self.hyper['MODEL']['NAME'] == 'FinalNet':
            self.model =  getattr(model, self.hyper['MODEL']['NAME'])(self.hyper).to(self.device)
        elif self.hyper['MODEL']['NAME'] == 'OneNet':
            self.model =  getattr(model, self.hyper['MODEL']['NAME'])().to(self.device)
        self.load()  # 加载日志目录下最好的模型
        self.model.eval()  # 切换到验证模式
        self.info = self.hyper['INFO']
        if AUTO_CALL:
            self.describe = self.hyper['DESCRIBE']  # EXP_NAME
            self.file_type = self.hyper['TEST']['FILE_TYPE']
            self.camera = self.hyper['TEST']['CAMERA']
            self.data_path = os.path.expanduser(self.hyper['TRAIN']['ROOT_PATH'])  # 数据集  '~/work/python/Public_DataSets/LSMI/' + self.file_type + '/' + self.camera + '_256/'
            self.result_path = self.hyper['TEST']['RESULT_PATH']+self.describe+'/'
            self.check_result_point_path = self.log_path + '/check_result_point/' + self.camera + '/'
            self.dataset = self.hyper['TEST']['DATASET']  # 在哪个数据集上测试
            self.single_ANGLE_path = self.log_path + '/single_MAE.xlsx'
            self.single_MSE_path = self.log_path + '/single_MSE.xlsx'
            self.ANGLE_path = self.result_path + 'MAE.xlsx'
            self.MSE_path = self.result_path + 'MSE.xlsx'
        else:
            self.describe = EXP_NAME  # EXP_NAME
            self.file_type = FILE_TYPE
            self.camera = CAMERA
            self.data_path = DATA_PATH
            self.result_path = RESULT_PATH + self.describe+'/'
            self.check_result_point_path = self.log_path + '/check_result_point/' + self.camera + '/'
            self.dataset = DATASET  # 在哪个数据集上测试
            self.single_ANGLE_path = self.log_path + '/single_MAE.xlsx'
            self.single_MSE_path = self.log_path + '/single_MSE.xlsx'
            self.ANGLE_path = self.result_path+ 'MAE.xlsx'
            self.MSE_path = self.result_path + 'MSE.xlsx'
        os.makedirs(self.result_path, exist_ok=True)
        os.makedirs(self.check_result_point_path, exist_ok=True)

    def test(self):
        print('-------------- 执行测试 --------------')
        torch.set_grad_enabled(False)  # 禁用梯度计算

        # 获取去除扩展名后的文件名前缀集合
        file_name = [os.path.splitext(f)[0] for f in os.listdir(self.data_path + self.dataset) if f.endswith('.' + self.file_type) and 'gt' not in f]  # 以偏色png图片名字为列表
        Angle_list = []
        MSE_list = []
        time_list = []
        for filename_prefix in tqdm(file_name):  # 遍历相同前缀的PNG和TIFF文件对
            # part = filename_prefix.split('_')

            # if len(part[1]) != 1:  # 只要单光源
            #     continue

            # if len(part[1])==1 or len(part[1])==3:  # 放弃单光源和3光源(只要2光源)
            #     continue

            # if len(part[1]) != 3:  # 只要3光源
            #     continue

            # stop = 0

            # 获取路径
            img_path = os.path.join(self.data_path, self.dataset, filename_prefix + '.' + self.file_type)
            # img_file_name = os.path.splitext(self.file_list[idx])[0]
            # gt_image_path = os.path.join(self.data_path, self.dataset, filename_prefix + '_gt.'+self.file_type)
            # gt_image_file_name = os.path.splitext(self.gt_image_list[idx])[0]
            mask_path = os.path.join(self.data_path, self.dataset, filename_prefix + '_mask.png')
            gt_path = os.path.join(self.data_path, 'illu_map', filename_prefix + ".npy")

            # 加载数据
            image = cv2.cvtColor(cv2.imread(img_path, cv2.IMREAD_UNCHANGED), cv2.COLOR_BGR2RGB).astype('float32')
            mask = cv2.cvtColor(cv2.imread(mask_path, cv2.IMREAD_UNCHANGED), cv2.COLOR_BGR2RGB).astype('float32')
            # gt_image = cv2.cvtColor(cv2.imread(gt_image_path, cv2.IMREAD_UNCHANGED), cv2.COLOR_BGR2RGB).astype('float32')
            gt = np.load(gt_path)
            # 转为张量，送入模型
            image_tensor = torch.from_numpy(image)
            image_tensor = image_tensor.permute(2, 0, 1)  # 转成3，255，255
            image_tensor = image_tensor / (image_tensor.max() + eps)
            image_tensor = image_tensor.unsqueeze(dim=0).to(self.device)

            start = time.perf_counter()
            # pred= model(image_tensor)
            # pred, pixel_illu, patch_illu, overall_illu, weight_map = self.model(image_tensor)
            pred, illu_map_all = self.model(image_tensor)
            end = time.perf_counter()
            elapsed = end - start
            time_list.append(elapsed)

            # ======================将光照图取平均并广播====================
            # if self.hyper['TRAIN']['ONLY_SINGLE_ILLU']:  # 纯单光源场景
            #     illu = pred.detach()
            #     illu = illu.squeeze()
            #     illu = illu.permute(1, 2, 0).cpu().numpy()
            #     mean_values = np.mean(illu, axis=(0, 1), keepdims=True)  # 计算平均值，得到 (1, 1, 3) keepdims=True 保持维度
            #     pred = np.broadcast_to(mean_values, illu.shape)  # 广播扩展回 (256, 256, 3)
            # ============================================================

            # # 检查中间光照图
            # # if filename_prefix == 'Place1013_123':  # 1013_123(val)
            # illu_map_all_0_tensor = illu_map_all[0] * 255
            # illu_map_all_0_tensor = illu_map_all_0_tensor.squeeze()
            # illu_map_all_0_tensor = illu_map_all_0_tensor.permute(1, 2, 0)
            # illu_map_all_0_np = illu_map_all_0_tensor.cpu().numpy().astype(np.uint8)
            # illu_map_all_0 = Image.fromarray(illu_map_all_0_np)
            # illu_map_all_0.save(self.check_result_point_path + filename_prefix + '_illu_map_all_0_image.png')
            #
            # illu_map_all_1_tensor = illu_map_all[1] * 255
            # illu_map_all_1_tensor = illu_map_all_1_tensor.squeeze()
            # illu_map_all_1_tensor = illu_map_all_1_tensor.permute(1, 2, 0)
            # illu_map_all_1_np = illu_map_all_1_tensor.cpu().numpy().astype(np.uint8)
            # illu_map_all_1_image = Image.fromarray(illu_map_all_1_np)
            # illu_map_all_1_image.save(self.check_result_point_path + filename_prefix + '_illu_map_all_1_image.png')
            #
            # # overall_illu_tensor = overall_illu * 255
            # # overall_illu_tensor = overall_illu_tensor.squeeze()
            # # overall_illu_tensor = overall_illu_tensor.permute(1, 2, 0)
            # # overall_illu_np = overall_illu_tensor.cpu().numpy().astype(np.uint8)
            # # overall_illu_image = Image.fromarray(overall_illu_np)
            # # overall_illu_image.save('./Result/L4_C512/' + filename_prefix + '_overall_illu_image.png')
            # #
            # # weight_map_tensor = weight_map * 255
            # # weight_map_tensor = weight_map_tensor.squeeze()
            # # weight_map_tensor = weight_map_tensor.permute(1, 2, 0)
            # # weight_map_np = weight_map_tensor.cpu().numpy().astype(np.uint8)
            # # weight_map_image = Image.fromarray(weight_map_np)
            # # weight_map_image.save('./Result/L4_C512/' + filename_prefix + '_weight_map_image.png')
            #
            # output_tensor = pred * 255
            # output_tensor = output_tensor.squeeze()
            # output_tensor = output_tensor.permute(1, 2, 0)
            # output_np = output_tensor.cpu().numpy().astype(np.uint8)
            # output_image = Image.fromarray(output_np)
            # output_image.save(self.check_result_point_path + filename_prefix + '_output_image.png')
            #
            # gt_tensor = gt * 255
            # # output_tensor = output_tensor.squeeze()
            # # output_tensor = output_tensor.permute(1, 2, 0)
            # gt_np = gt_tensor.astype(np.uint8)
            # gt_image = Image.fromarray(gt_np)
            # gt_image.save(self.check_result_point_path + filename_prefix + '_gt_image.png')

            pred = pred.squeeze().permute(1, 2, 0)  # 转为numpy数组
            pred = pred.cpu().numpy()
            pred = pred * mask  # 去掉色卡

            # 计算MSE
            # MSE = np.mean((gt - pred) ** 2)
            # MSE_list.append(MSE)

            pred = pred.reshape(pred.shape[0] * pred.shape[1], 3)
            gt = gt.reshape(gt.shape[0] * gt.shape[1], 3)

            # pred_mean = np.mean(pred, axis=0)
            # gt_mean = np.mean(gt, axis=0)
            # STOP = 0
            mask = mask.reshape(mask.shape[0] * mask.shape[1], 3)
            row_mask = np.any(mask != 0, axis=1)
            # 使用布尔索引剔除 mask 值为 0 的行
            filtered_output = pred[row_mask]
            filtered_gt = gt[row_mask]

            # 计算每张图片的MSE（过滤mask）
            a_pic_MSE = np.mean((filtered_gt - filtered_output) ** 2)
            MSE_list.append(a_pic_MSE)

            # 计算每张图片的角度差（过滤mask）
            angle = getMAE_numpy_2D(filtered_output, filtered_gt)
            a_pic_mean = angle.mean()

            # # 计算每张图片的MSE(单光源，单像素)
            # MSE = np.mean((gt_mean - pred_mean) ** 2)
            # MSE_list.append(MSE)
            #
            # # 计算每张图片的角度差（单光源，单像素）
            # angle = arccos_numpy(gt_mean, pred_mean)
            # # a_pic_mean = angle.mean()




            # =================== 记录每张光照图的角度差(MAE)，并统计 ========================
            if not os.path.exists(self.single_ANGLE_path):
                wb = Workbook()
                ws = wb.active
                ws.title = 'Sheet1'
                ws.append(['Name', 'Angle'])  # 写入表头
                wb.save(self.single_ANGLE_path)

            df = pd.read_excel(self.single_ANGLE_path)  # 读取已有的 Excel 文件
            new_data = {'Name': [filename_prefix], 'Angle': [a_pic_mean]}  # 创建新的数据行
            new_df = pd.DataFrame(new_data)
            df = pd.concat([df, new_df], ignore_index=True)  # 将新数据行追加到原有的 DataFrame 中
            df.to_excel(self.single_ANGLE_path, index=False)  # 将更新后的 DataFrame 写回到 Excel 文件中
            # =====================================================================

            # =================== 记录每张光照图的均方差(MSE)，并统计 ========================
            if not os.path.exists(self.single_MSE_path):
                wb = Workbook()
                ws = wb.active
                ws.title = 'Sheet1'
                ws.append(['Name', 'Angle'])  # 写入表头
                wb.save(self.single_MSE_path)

            df = pd.read_excel(self.single_MSE_path)  # 读取已有的 Excel 文件
            new_data = {'Name': [filename_prefix], 'Angle': [a_pic_MSE]}  # 创建新的数据行
            new_df = pd.DataFrame(new_data)
            df = pd.concat([df, new_df], ignore_index=True)  # 将新数据行追加到原有的 DataFrame 中
            df.to_excel(self.single_MSE_path, index=False)  # 将更新后的 DataFrame 写回到 Excel 文件中
            # =====================================================================

            Angle_list.append(a_pic_mean)
        # ================= fps ====================
        total_time = sum(time_list)
        total_number = len(time_list)
        # avg = total_time / total_number
        fps = total_number / total_time
        # print('avg:', avg)
        print('fps:', fps)
        # =========================================

        def statistic(alist, save_path):
            alist = sorted(alist)
            total = sum(alist)
            n = len(alist)
            avg = total / n

            # 计算每个数据点与平均值的差的平方，并求和
            squared_diffs = sum((x - avg) ** 2 for x in alist)
            # 计算方差（标准差的平方）
            variance = squared_diffs / len(alist)
            # 计算标准差
            std = variance ** 0.5

            if n % 2 == 1:
                median = alist[n // 2]
            else:
                median = (alist[n // 2 - 1] + alist[n // 2]) / 2

            # 四分位数
            q1_index = int(n * 0.25)
            q3_index = int(n * 0.75)
            q1 = alist[q1_index]
            q3 = alist[q3_index]

            # 3均值
            trimean = (2 * median + q1 + q3) / 4

            # 前25%的平均值
            best25_avg = sum(alist[:q1_index + 1]) / (q1_index + 1)  # +1 to include the q1 value if needed

            # 后75%的平均值 (from q1 to the end, inclusive of q1)
            last_75_percent = alist[q1_index:]
            worst25_avg = sum(last_75_percent) / len(last_75_percent)

            # 第95%位的值
            ninety_fifth_percentile_index = int((n - 1) * 0.95 + 0.5)  # Rounding to nearest integer
            per95 = alist[ninety_fifth_percentile_index]
            print("Mean:", avg,
                  "\nStd:", std,
                  "\nmedian:", median,
                  "\ntrimean:", trimean,
                  "\nbest25:", best25_avg,
                  "\nworst25:", worst25_avg,
                  "\npercentile95:", per95)
            # ===================保存excel========================
            if not os.path.exists(save_path):
                wb = Workbook()
                ws = wb.active
                ws.title = 'Sheet1'
                ws.append(['Name', 'Mean', 'Std', 'Median', 'Trimean', 'Best25', 'Worst25', '95Per', '备注'])  # 写入表头
                wb.save(save_path)

            # 加载现有的工作簿和工作表
            wb = load_workbook(save_path)
            ws = wb['Sheet1']
            ws.append([self.camera + '>>>' + self.describe + '>>>' + self.dataset,
                       avg, std, median, trimean, best25_avg, worst25_avg, per95, self.info+'>>>'+self.log_path+'>帧率：'+str(fps)])
            wb.save(save_path)
            # ===================================================
            print(f'新数据已追加到 {save_path}\n')
            return 0

        statistic(Angle_list, self.ANGLE_path)  # 统计角度差
        statistic(MSE_list, self.MSE_path)  # 统计均方差
        print('-------------- 测试完毕 --------------')

    # 加载模型
    def load(self):
            check_path = os.path.join(self.log_path, 'model_best.pt')
            check_point = torch.load(check_path)
            self.model.load_state_dict(check_point['model'])


if __name__ == '__main__':
    AUTO_CALL = False  # 手动调用测试
    # =============================== Hyper Parameter ===================================
    EXP_NAME = 'Compare'  # 实验名称
    CAMERA = 'galaxy'
    FILE_TYPE = 'tiff'  # tiff 还是 png
    DATASET = 'test'  # 用哪种数据集测试
    MODEL_PATH = './Result/' + EXP_NAME + '/09300250/'  # 训练好的模型路径
    DATA_PATH = os.path.expanduser('~/work/python/Public_DataSets/LSMI/' + FILE_TYPE + '/' + CAMERA + '_256/')  # 数据集
    RESULT_PATH = './Result/'
    # single_ANGLE_path = MODEL_PATH+ 'single_check_result.xlsx'
    # ANGLE_path = RESULT_PATH + 'check_result.xlsx'
    # MSE_path = RESULT_PATH + 'check_result_MSE.xlsx'
    # ===================================================================================
    tester = Tester(MODEL_PATH, config_file_name = 'hyper_parameters2.yaml')
    tester.test()
